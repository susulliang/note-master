"""Convert GBU XLSX sheets to per-sheet GFM Markdown files in products/."""
import openpyxl, re, os
from pathlib import Path

XLSX = "/workspace/.uploads/c2f74b91-53eb-4314-85e1-788c7a52d714_史上最全GBU (北美).xlsx"
OUT_DIR = Path("/workspace/products")
OUT_DIR.mkdir(exist_ok=True)

NAV_PATTERNS = [
    re.compile(r"^点击.*返回.*主页"),
    re.compile(r"^点击超链接自动跳转"),
]
SKIP_VALUES = {"", "nan", "none", "NaN", "NaT"}

def cell_str(v):
    if v is None:
        return ""
    s = str(v).replace("\u00a0", " ").strip()
    for pat in NAV_PATTERNS:
        if pat.search(s):
            return ""
    return "" if s in SKIP_VALUES else s


def propagate_merges(ws):
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[""] * (max_c + 1) for _ in range(max_r + 1)]
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            grid[r][c] = cell_str(ws.cell(row=r, column=c).value)
    for mr in ws.merged_cells.ranges:
        val = grid[mr.min_row][mr.min_col]
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if grid[r][c] == "":
                    grid[r][c] = val
    return grid, max_r, max_c


def find_rightmost_col(grid, max_r, max_c):
    last = 0
    for c in range(1, max_c + 1):
        if any(grid[r][c] != "" for r in range(1, max_r + 1)):
            last = c
    return last


def find_bottom_row(grid, max_r, cols):
    last = 0
    for r in range(1, max_r + 1):
        if any(grid[r][c] != "" for c in cols):
            last = r
    return last


def is_super_title(row_vals):
    nonempty = [v for v in row_vals if v != ""]
    if 1 <= len(nonempty) <= 2 and len(row_vals) >= 6:
        joined = " ".join(nonempty)
        if re.search(r"(?i)\b(SPEC|version|rev\.?|v\d)", joined):
            return True
    return False


def sectionize(grid, rows, cols):
    sections = []
    cur_cat, cur_rows = None, []
    for r in rows:
        cat = grid[r][cols[0]]
        if cat == "":
            cur_rows.append(r)
            continue
        if cat != cur_cat:
            if cur_rows:
                sections.append((cur_cat, cur_rows))
            cur_cat, cur_rows = cat, [r]
        else:
            cur_rows.append(r)
    if cur_rows:
        sections.append((cur_cat, cur_rows))
    return [(c, rs) for c, rs in sections if rs]


def detect_model_labels(grid, rows, cols):
    model_cols = cols[2:]
    labels = {c: "" for c in model_cols}
    for c in model_cols:
        for r in rows:
            if grid[r][c] != "":
                labels[c] = grid[r][c]
                break
    # Replace with Marketing Name if such a row exists
    for r in rows:
        field = grid[r][cols[1]]
        if re.match(r"(?i)marketing\s*name", field):
            for c in model_cols:
                v = grid[r][c]
                if v != "":
                    bom = labels[c]
                    if bom and bom.lower() != v.lower() and bom not in v:
                        labels[c] = f"{v} ({bom})" if len(bom) < 14 else v
                    else:
                        labels[c] = v
            break
    return labels


def sanitize(s):
    s = (s or "").replace("\\", "\\\\").replace("|", "\\|")
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    return s.strip()


SAFE = re.compile(r"[^a-zA-Z0-9_\-]+")


def filename(i, name):
    stem = name.strip().replace("&", "and").replace(" ", "_").replace("/", "-")
    stem = SAFE.sub("_", stem).strip("_")
    num = 10 + i - 1
    return f"{num:02d}_{stem}.md"


def process_sheet(idx, name, ws):
    grid, max_r, max_c = propagate_merges(ws)
    last_c = find_rightmost_col(grid, max_r, max_c)
    if last_c < 3:
        return None
    cols = list(range(1, last_c + 1))
    last_r = find_bottom_row(grid, max_r, cols)
    if last_r < 2:
        return None
    rows_all = list(range(1, last_r + 1))
    super_titles = []
    cursor = 1
    while cursor <= last_r:
        vals = [grid[cursor][c] for c in cols]
        if is_super_title(vals):
            super_titles.append(" / ".join(v for v in vals if v != ""))
            cursor += 1
        else:
            break
    rows = list(range(cursor, last_r + 1))
    if not rows:
        return None
    sections = sectionize(grid, rows, cols)
    if not sections:
        return None
    labels = detect_model_labels(grid, rows, cols)
    model_cols = cols[2:]
    if not model_cols:
        return None

    md = [f"## {name}", ""]
    meta = [f"Source: GBU (北美)", f"Sheet: {name}", f"Rows×Cols: {last_r}×{len(cols)}"] + super_titles
    md.append("*" + " · ".join(meta) + "*")
    md.append("")
    mnames = [labels[c] for c in model_cols if labels.get(c, "")]
    if mnames:
        md.append("**Models covered:** " + ", ".join(f"`{n}`" for n in mnames))
        md.append("")
    for cat, srows in sections:
        cat_display = (cat or "General").strip()
        kept = [r for r in srows if any(grid[r][c] != "" for c in cols[1:])]
        if not kept:
            continue
        md.append(f"### {cat_display}")
        md.append("")
        headers = ["Spec"] + [labels.get(c, f"Col{c}") for c in model_cols]
        md.append("| " + " | ".join(sanitize(h) for h in headers) + " |")
        md.append("| " + " | ".join("---" for _ in headers) + " |")
        for r in kept:
            spec = grid[r][cols[1]]
            values = [grid[r][c] for c in model_cols]
            if spec == "" and all(v == "" for v in values):
                continue
            cells = [sanitize(spec)] + [sanitize(v) for v in values]
            md.append("| " + " | ".join(cells) + " |")
        md.append("")
    return "\n".join(md).rstrip() + "\n"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheet_names = wb.sheetnames
    written = 0
    skipped = []
    for i, name in enumerate(sheet_names, 1):
        ws = wb[name]
        content = process_sheet(i, name, ws)
        if not content:
            skipped.append(name)
            continue
        p = OUT_DIR / filename(i, name)
        if p.exists():
            p = OUT_DIR / filename(i, f"GBU_{name}")
        with open(p, "w", encoding="utf-8") as f:
            f.write(content)
        written += 1
    print(f"Written {written} GBU markdown files to {OUT_DIR}")
    if skipped:
        print(f"Skipped: {skipped}")

    readme = OUT_DIR / "README.md"
    lines = readme.read_text(encoding="utf-8").splitlines() if readme.exists() else []
    cut = None
    for i, l in enumerate(lines):
        if l.startswith("## GBU (北美) — Sheet Index"):
            cut = i
            break
    if cut is not None:
        lines = lines[:cut]
    while lines and lines[-1].strip() == "":
        lines.pop()
    lines += [
        "",
        "## GBU (北美) — Sheet Index",
        "",
        f"Source: `史上最全GBU (北美).xlsx` — {len(sheet_names)} sheets.",
        "",
        "| # | Sheet | Output | Rows×Cols | Notes |",
        "|---|---|---|---|---|",
    ]
    notes_map = {
        "Barcode": "SKU / model cross-reference + navigation index",
        "AES": "Auto-Empty Station accessory",
        "Deebot": "Legacy Deebot spec comparison (many models)",
        "OZMO": "OZMO legacy lineup comparison",
    }
    for i, name in enumerate(sheet_names, 1):
        ws = wb[name]
        fn = filename(i, name)
        if not (OUT_DIR / fn).exists():
            fn = filename(i, f"GBU_{name}")
        note = ""
        if name in notes_map:
            note = notes_map[name]
        elif name.startswith("EOLed"):
            note = "Discontinued / EOL models"
        elif any(name.lower().startswith(p) for p in ["x", "t", "n", "w", "goat", "gx"]):
            note = "Series spec comparison (multiple variants)"
        lines.append(
            f"| {i} | {name} | [`{fn}`](./{fn}) | {ws.max_row}×{ws.max_column} | {note} |"
        )
    lines.append("")
    readme.write_text("\n".join(lines), encoding="utf-8")
    print("Updated products/README.md with GBU index.")


if __name__ == "__main__":
    main()
